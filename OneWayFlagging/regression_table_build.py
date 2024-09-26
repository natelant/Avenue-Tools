# all the import libraries here
import os
import xml.etree.ElementTree as ET
import math
from datetime import datetime
from openpyxl import load_workbook
import gpxpy
from geopy import distance
import pandas as pd
import numpy as np
import pytz
from fastkml import kml
import zipfile
# INPUTS... folder name or raw files... maybe this will just loop through the folders in the folder path...
# needs two counts folders (one in each direction), GPX, KML, and a csv/txt with the route info (speed limit, multiple access, pilot car)
# KML needs to have the direction in the name





# definitions - note we are evaluating direction separately. Opposite direction information (volumes, green time, red time, all red time) will be evaluated in interactions 

# Define Salt Lake City's timezone
LOCAL_TZ = pytz.timezone('America/Denver')

def read_files(folder_path):
    files = {}
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if filename.endswith('.gpx'):
            files['gpx'] = file_path
        elif filename.endswith('.kml'):
            files['kml'] = file_path
        elif filename.endswith('.xlsm'):
            if 'xlsm1' not in files:
                files['xlsm1'] = file_path
            else:
                files['xlsm2'] = file_path
    return files

def parse_xlsm(xlsm_file):
    wb = load_workbook(xlsm_file, read_only=True, keep_links=False)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=1, max_col=3, values_only=True):
        data.append(row)
    return data

# Haversine formula to calculate distance between two points on Earth
def haversine(lat1, lon1, lat2, lon2):
    R = 6371000  # Earth radius in meters
    
    dlat = np.radians(lat2 - lat1)
    dlon = np.radians(lon2 - lon1)
    
    a = np.sin(dlat / 2) ** 2 + np.cos(np.radians(lat1)) * np.cos(np.radians(lat2)) * np.sin(dlon / 2) ** 2
    c = 2 * np.arctan2(np.sqrt(a), np.sqrt(1 - a))
    
    distance = R * c  # in meters
    return distance

# Function to calculate the bearing between two points
def calculate_bearing(lat1, lon1, lat2, lon2):
    lat1 = math.radians(lat1)
    lon1 = math.radians(lon1)
    lat2 = math.radians(lat2)
    lon2 = math.radians(lon2)

    dlon = lon2 - lon1
    x = math.sin(dlon) * math.cos(lat2)
    y = math.cos(lat1) * math.sin(lat2) - (math.sin(lat1) * math.cos(lat2) * math.cos(dlon))
    initial_bearing = math.atan2(x, y)

    # Convert bearing from radians to degrees
    initial_bearing = math.degrees(initial_bearing)
    compass_bearing = (initial_bearing + 360) % 360

    return compass_bearing

# Function to convert bearing to compass direction
def bearing_to_direction(bearing):
    if 45 <= bearing < 135:
        return 'EB'
    elif 135 <= bearing < 225:
        return 'SB'
    elif 225 <= bearing < 315:
        return 'WB'
    else:
        return 'NB'

# Function to parse GPX file and calculate speed (and grade) returns a df
def parse_gpx_file(gpx_file, intersections, route_name):
    gpx_data = []
    # Parse the GPX file
    with open(gpx_file, 'r') as gpx_file:
        gpx = gpxpy.parse(gpx_file)
        # Extract data from GPX
        for track in gpx.tracks:
            for segment in track.segments:
                for point in segment.points:
                    # Convert UTC time to local time
                    local_time = point.time.replace(tzinfo=pytz.UTC).astimezone(LOCAL_TZ)
                    gpx_data.append({
                        'lat': point.latitude, 
                        'lon': point.longitude, 
                        'time': local_time.strftime('%Y-%m-%dT%H:%M:%S'),
                        'elevation': point.elevation
                        })

    # Initialize variables
    output_data = []  # Initialize output data list
    prev_intersection = None  # Initialize previous intersection

    # Iterate through each point in GPX data
    for point in gpx_data:
        closest_intersection = None
        # Iterate through each significant intersection
        for intersection in intersections:
            # Calculate the distance between the current point and the intersection
            distance = haversine(point['lat'], point['lon'], intersection['lat'], intersection['lon'])
            # Check if the distance is within the threshold (50 feet)
            if distance <= 30:  # 100 feet in meters
                closest_intersection = intersection
                # Assign the time and elevation of the current point to the closest intersection
                closest_intersection['time'] = point['time']
                closest_intersection['elevation'] = point['elevation']
                break  # Exit the loop if a significant intersection is found within the threshold

        # If a closest intersection is found within the threshold, process it
        if closest_intersection:
            if prev_intersection:
                # Calculate total distance between the previous intersection and the current closest intersection
                total_distance = haversine(prev_intersection['lat'], prev_intersection['lon'], closest_intersection['lat'], closest_intersection['lon'])
                
                # Calculate the bearing between the previous intersection and the current closest intersection
                bearing = calculate_bearing(prev_intersection['lat'], prev_intersection['lon'], closest_intersection['lat'], closest_intersection['lon'])
                
                # Convert the bearing to a compass direction
                direction = bearing_to_direction(bearing)
                
                # Calculate time difference in seconds
                time_diff = (datetime.strptime(closest_intersection['time'], '%Y-%m-%dT%H:%M:%S') - datetime.strptime(prev_intersection['time'], '%Y-%m-%dT%H:%M:%S')).total_seconds()
                
                # Calculate average speed (m/s) and convert to mph
                avg_speed = (total_distance / time_diff) * 2.23694 if time_diff != 0 else 0
                
                # Calculate elevation difference
                elevation_diff = closest_intersection['elevation'] - prev_intersection['elevation']
                
                # Calculate average grade
                avg_grade = (elevation_diff / total_distance) * 100 if total_distance != 0 else 0
                
                # Store the segment start, segment finish, average speed, average grade, total distance, and direction in the output data
                output_data.append({
                    'route_name': route_name,
                    'segment_ID': prev_intersection['pin_id'] + ' / ' + closest_intersection['pin_id'], 
                    'segment_start': prev_intersection['segment_id'], 
                    'segment_finish': closest_intersection['segment_id'], 
                    'avg_speed': avg_speed, # mph
                    'avg_grade': avg_grade, # percent
                    'total_distance': total_distance * 0.000621371, # convert to miles
                    'start_time': prev_intersection['time'], 
                    'end_time': closest_intersection['time'],
                    'Direction': direction})  # Add direction to the output data
            prev_intersection = closest_intersection

    # Convert output data to DataFrame
    output_df = pd.DataFrame(output_data)
    # filter out any rows where segment_start and segment_finish are the same
    output_df = output_df[output_df['segment_start'] != output_df['segment_finish']]
    # filter out any rows where segment_finish contains 'Queue' or 'queue'
    output_df = output_df[~output_df['segment_finish'].str.contains('Queue|queue', case=False)].reset_index(drop=True)
    # group by route_ID and calculate average speed, grade, and distance
    output_df = output_df.groupby(['route_name', 'segment_ID', 'Direction']).agg({
        'avg_speed': 'mean',
        'avg_grade': 'mean',
        'total_distance': 'mean'
    }).reset_index()
    # create new column for segment_type where if segment_ID contains 'Queue' or 'queue' then segment_type = 'Queue' else segment_type = 'Segment'
    output_df['segment_type'] = output_df['segment_ID'].str.contains('Queue|queue', case=False).apply(lambda x: 'Queue' if x else 'Segment').reset_index(drop=True)
    # group by route_id and direction and then create column for average speed in queue, average speed in segment, average grade (weighted by distance), total distance in segment
    grouped_df = output_df.groupby(['route_name', 'Direction'])

    avg_speed_queue = grouped_df.apply(lambda x: x[x['segment_type'] == 'Queue']['avg_speed']).reset_index(name='avg_speed_queue')
    avg_speed_segment = grouped_df.apply(lambda x: x[x['segment_type'] == 'Segment']['avg_speed']).reset_index(name='avg_speed_segment')

    # Calculate average grade (weighted by distance)
    avg_grade_weighted = grouped_df.apply(lambda x: np.average(x['avg_grade'], weights=x['total_distance'])).reset_index(name='avg_grade_weighted')

    # Calculate total distance in segment
    total_distance_segment = grouped_df.apply(lambda x: x[x['segment_type'] == 'Segment']['total_distance']).reset_index(name='total_distance_segment')

    # Merge all the calculated columns into a single DataFrame
    final_df = avg_speed_queue.merge(avg_speed_segment, on=['route_name', 'Direction'])
    final_df = final_df.merge(avg_grade_weighted, on=['route_name', 'Direction'])
    final_df = final_df.merge(total_distance_segment, on=['route_name', 'Direction'])

    # if the dataframe contains columns with 'level' in the name, drop them
    final_df = final_df[[col for col in final_df.columns if 'level' not in col]]

    return final_df



# Function to convert KMZ to KML and return the path to the KML file
def convert_kmz_to_kml(kmz_file):
    with zipfile.ZipFile(kmz_file, 'r') as kmz:
        for file_name in kmz.namelist():
            if file_name.endswith('.kml'):
                kmz.extract(file_name, os.path.dirname(kmz_file))
                return os.path.join(os.path.dirname(kmz_file), file_name)
    return None

# function to parse KML file containing significant intersection information

def parse_kml(folder):
    intersections = []

    # search folder for the kml file
    for filename in os.listdir(folder):
        if filename.lower().endswith('.kml'):
            file_path = os.path.join(folder, filename)
            print(f'Processing file: {filename}')
            kml_file = file_path
        elif filename.lower().endswith('.kmz'):
            # might need some QC here...
            kml_file = convert_kmz_to_kml(filename)

    
    with open(kml_file, 'rb') as file:
        doc = file.read()
        k = kml.KML()
        k.from_string(doc)
        
        # Assuming the KML has a single document
        for document in k.features():
            for folder in document.features():
                for feature in folder.features():
                    if isinstance(feature, kml.Placemark):
                        geometry = feature.geometry
                        if geometry:
                            if geometry.geom_type == 'Point':
                                coordinates = list(geometry.coords)[0]
                                intersections.append({
                                    'pin_id': feature.name,
                                    'segment_id': feature.name,
                                    #'description': feature.description if feature.description else '',
                                    'lat': coordinates[1],
                                    'lon': coordinates[0]
                                })
                            
    return intersections

# function that cleans the excel files
def clean_xlsm(filename, route_name):
    # Read the Excel file into a dataframe
    df = pd.read_excel(filename)
    # keep only the first 3 columns
    df = df.iloc[:, :3]
    
    # Filter out invalid segments between multiple 'Start's without an intervening 'Stop'
    segments = []
    start_idx = None
    valid_segment = True

    for i, row in df.iterrows():
        if row['Vehicle'] == 'Start':
            if start_idx is not None:
                # Exclude the segment if 'Start' is immediately followed by 'Stop'
                if i == start_idx + 1 and df.loc[i, 'Vehicle'] == 'Stop':
                    start_idx = None
                    valid_segment = False
                    continue
            start_idx = i
            valid_segment = True
        elif row['Vehicle'] == 'Stop':
            if start_idx is not None:
                # Include only segments with valid data between 'Start' and 'Stop'
                if valid_segment:
                    segments.append(df.loc[start_idx:i])
            start_idx = None
            valid_segment = False

    # Concatenate valid segments into a single DataFrame
    if segments:
        cleaned = pd.concat(segments)
    else:
        cleaned = pd.DataFrame(columns=df.columns)

    # remove consecutive start-stops
    drop_rows = cleaned[(cleaned['Vehicle'] == 'Start') & (cleaned['Vehicle'].shift(-1) == 'Stop')].index
    drop_rows = drop_rows.union(drop_rows+1)

    cleaned = cleaned.drop(drop_rows)

    #remove rows with nan in the Vehicle column
    cleaned = cleaned[cleaned['Vehicle'].notna()]
    # Vehicle should be a string
    cleaned['Vehicle'] = cleaned['Vehicle'].astype(str)

    # Add route_name column
    cleaned['route_name'] = route_name

    return cleaned


# write fucntion to convert Time Stamp to a datetime variable
def time_to_datetime(t):
    return datetime.combine(datetime.today(), t)

# generate counts from the xlsm files
def generate_counts(xlsm_files, route_name):
    # Initialize empty lists to hold data for counts and splits
    counts_data = []
    splits_data = []
    hourly_data = []
    
    for file in xlsm_files:
        # Read the XLSM file
        df = clean_xlsm(file, route_name)

        # Calculate headway here
        # Convert 'time stamp' to datetime for continuous axis plotting
        df['Timestamp'] = df['Time Stamp'].apply(time_to_datetime)
        df['headway'] = df.groupby(['route_name', 'Direction'])['Timestamp'].diff().dt.total_seconds()
        # Initialize variables
        split_id = 1
        split_data = []
        
        # Loop through rows and create split_id
        for _, row in df.iterrows():
            split_data.append(row.to_dict())
            split_data[-1]['split_id'] = split_id
            
            # Check if 'Vehicle' is a string and compare to 'stop'
            if isinstance(row['Vehicle'], str) and row['Vehicle'].lower() == 'stop':
                split_id += 1
        
        # Loop through the split_data list to separate counts and splits
        for row in split_data:
            if row['Vehicle'].lower() == 'start' or row['Vehicle'].lower() == 'stop':
                # Append to splits_data if the vehicle is 'Start' or 'Stop'
                splits_data.append(row)
            else:
                # Change 'Probe' or 'probe' to 'Car' in the 'Vehicle' column
                if row['Vehicle'].lower() == 'probe':
                    row['Vehicle'] = 'Car'
                # Append to counts_data if the vehicle is not 'Start' or 'Stop'
                counts_data.append(row)
                hourly_data.append(row)

    # Group counts_data by route_name, split_id, and direction
    grouped_counts = pd.DataFrame(counts_data).groupby(['route_name', 'split_id', 'Direction'])
    
    # Summarize the sum of vehicles, the average headway, and calculate the truck percentage for each group
    counts_df = grouped_counts.agg({
        'Vehicle': ['count', lambda x: (x == 'Truck').sum() / len(x) * 100],  # Total count and truck percentage
        'headway': 'mean'
    })
    
    # Rename the columns
    counts_df.columns = ['volume', 'truck_percentage', 'avg_headway']
    
    # Reset the index to make route_name, split_id, and Direction regular columns
    counts_df = counts_df.reset_index()
    
    # Convert lists to DataFrames for splits and hourly data
    splits_df = pd.DataFrame(splits_data)
    # Order the splits_df by 'Timestamp'
    splits_df = splits_df.sort_values(by='Timestamp')
    # group by route_name and direction and pivot by Vehicle and fill with Timestamp
    splits_df = splits_df.pivot(index=['route_name', 'Direction', 'split_id'], columns='Vehicle', values='Timestamp')
    # (grouped by route_name and direction) calculate split time, green time, and red time
    # green time is stop time - start time
    # split time is start time - previous start time
    # red time is split time - green time
    # all red time is start time - previous stop time
    splits_df['green_time'] = (splits_df['Stop'] - splits_df['Start']).dt.total_seconds()
    splits_df['split_time'] = (splits_df['Start'] - splits_df.groupby(level=['route_name', 'Direction'])['Start'].shift(1)).dt.total_seconds()
    splits_df['red_time'] = splits_df['split_time'] - splits_df['green_time']
    splits_df['all_red_time'] = (splits_df['Start'] - splits_df.groupby(level=['route_name', 'Direction'])['Stop'].shift(1)).dt.total_seconds()
    
    # filter out rows where split_time is greater than 1000
    splits_df = splits_df[(splits_df['split_time'] < 1000) | (splits_df['split_time'].isna())]

    hourly_df = pd.DataFrame(hourly_data)

    # Group by route_name and Direction
    grouped = hourly_df.groupby(['route_name', 'Direction'])
    
    # Calculate total volume and total time for each group
    hourly_summary = grouped.agg({
        'Vehicle': 'count',  # Total volume
        'Timestamp': lambda x: (x.max() - x.min()).total_seconds() / 3600  # Total time in hours
    }).reset_index()
    
    # Rename columns
    hourly_summary.columns = ['route_name', 'Direction', 'Total_Volume', 'Total_Hours']
    
    # Calculate volume per hour
    hourly_summary['Volume_Per_Hour'] = hourly_summary['Total_Volume'] / hourly_summary['Total_Hours']
    # drop the Total_Hours and total volume columns
    hourly_summary = hourly_summary.drop(columns=['Total_Hours', 'Total_Volume'])
        
        
    # counts is a combined df of both count xlsm files the contains volume headway and truck percentage for each split
    # splits is a combined df of both count xlsm files and contains the split_id, green_time, red_time, and start_time
    # hourly is a combined df of both count xlsm files and contains the total hourly volume for each direction
    return counts_df, splits_df, hourly_summary




# Usage example:
# gpx_df = parse_gpx(files['gpx'])
# kml_intersections = parse_kml(folder_path)
# segment_df = create_segment_dataframe(gpx_df, kml_intersections)
# print(segment_df)





## calculate hourly volume from the counts

## calculates splits, volumes, and headways from counts






#!!!# Loop through each folder, join tables, append to overall data table (splits)
# test that the data is clean (start, stop, etc. not repeating start, start, etc.) and give warning that the data needs to be reviewed. Or warn that directions were skipped...?

folder_path = "OneWayFlagging/regression_input/US-6"
files = read_files(folder_path)
route_name = os.path.basename(folder_path)
print('route_name:', route_name)

# parse the kml file and get the intersections
kml_intersections = parse_kml(folder_path)

# parse the gpx file and get the segment information
gpx_df = parse_gpx_file(files['gpx'], kml_intersections, route_name) # returns a df with segment_ID, avg_speed, avg_grade, total_distance

# parse the xlsm files and generate counts, splits, and volumes, etc.
xlsm_files = [files['xlsm1'], files['xlsm2']]
counts_df, splits_df, hourly_df = generate_counts(xlsm_files, route_name)


# Merge gpx_df, counts_df, and hourly_df into splits_df by route_name and Direction
merged_df = pd.merge(splits_df.reset_index(), gpx_df, on=['route_name', 'Direction'], how='left')
merged_df = pd.merge(merged_df, counts_df, on=['route_name', 'Direction', 'split_id'], how='left')
merged_df = pd.merge(merged_df, hourly_df, on=['route_name', 'Direction'], how='left')
print(merged_df.head(60))


# -----------------------------------------------------------------------------------------------------------------------------------------------
# interactions - opposite direction information - additional variables and possible calculations like saturated flow rates, capacity, delay, etc.